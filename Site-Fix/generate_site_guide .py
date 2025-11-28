#!/usr/bin/env python3
"""
Website Revision Guide Generator v3.0
サイト構成ガイド自動生成ツール

【出力の粒度ルール】
- ページ名：トップページ、記事一覧ページ、このメディアについて、お問い合わせページなど
- セクション名：ヘッダー、ヒーローセクション、記事一覧、フッターなど
- 内容：【セクション名】形式で始まり、すべてのテキスト要素を改行区切りで記載
  - 見出しテキスト
  - 説明文・本文
  - ボタン・リンクテキスト（►形式）
  - 記事のタイトル・カテゴリ・日付・著者・リード文
  - フォームの項目名
  - 注記・補足
"""

import os
import yaml
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_guide(config_path='config.yaml', output_dir='output'):
    """サイト構成ガイドを生成"""
    
    # 設定ファイルを読み込む
    with open(config_path, 'r', encoding='utf-8') as f:
        config = yaml.safe_load(f)
    
    # 新しいワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = "サイト構成ガイド"
    ws.sheet_view.showGridLines = False
    
    # スタイル定義
    header_fill = PatternFill(
        start_color="9BC2E6",  # 明るい青3
        end_color="9BC2E6",
        fill_type="solid"
    )
    header_font = Font(color="000000", bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # タイトル行
    ws['A1'] = f'{config["site_name"]} - サイト構成ガイド'
    ws['A1'].font = Font(size=18, bold=True, color="000000")
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 30
    
    # 空白行
    ws.row_dimensions[2].height = 15
    
    # 修正指示の説明セクション
    instruction_text = """【重要】修正内容のご記入方法について"""
    ws['A3'] = instruction_text
    ws['A3'].font = Font(size=14, bold=True, color="000000")
    ws.merge_cells('A3:D3')
    
    instruction_detail = """修正内容を正しく認識し、スムーズに反映できるよう、
お手数ですが修正指示は以下の点をご記入くださいますようご協力をお願いいたします。

1. 修正箇所を具体的にご記入ください。
   例）「一番上のテキスト」「右側の説明文」「カードの2枚目」「見出しの下の本文」など

2. どのような修正を行うかを具体的にご記入ください。
   例）
   ・テキスト →「文言の変更」「追記」「削除」「位置を変更」
   ・ボタン →「文言の変更」「リンク先の変更」

3. 修正後の具体的な内容をご記入ください。
   例）
   ・テキスト →「『○○』の文を『□□』に変更」
   ・配置　→「本文の上に配置」「タイトルの下に移動」

※「この部分」「先日の内容」「いい感じに」「おまかせで」など、
  対象や内容が特定できない表現は、できるだけお控えください。
"""
    ws['A4'] = instruction_detail
    ws['A4'].font = Font(size=11, color="000000")
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A4:D4')
    ws.row_dimensions[4].height = 250
    
    # 空白行
    ws.row_dimensions[5].height = 15
    
    # ヘッダー行（6行目に配置）
    headers = ['ページ名', 'セクション名', '内容', '修正内容']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    current_row = 7
    
    # ページごとの開始行を記録
    page_start_rows = {}
    
    def add_section_row(page_name, section_name, content):
        """セクション行を追加"""
        nonlocal current_row
        
        # ページ名のセル統合処理
        if page_name not in page_start_rows:
            page_start_rows[page_name] = current_row
            ws.cell(row=current_row, column=1).value = page_name
        else:
            ws.cell(row=current_row, column=1).value = ""
        
        ws.cell(row=current_row, column=2).value = section_name
        ws.cell(row=current_row, column=3).value = content
        
        # スタイル適用
        for col_idx in range(1, 5):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            if col_idx == 1:
                cell.font = Font(bold=True, color="000000")
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
            elif col_idx == 2:
                cell.font = Font(bold=True, color="000000")
            else:
                cell.font = Font(color="000000")
        
        # 内容の量に応じて行の高さを調整
        line_count = content.count('\n') + 1
        ws.row_dimensions[current_row].height = max(30, min(line_count * 15, 600))
        
        current_row += 1
    
    # 共通部分からヘッダーとフッターを取得
    header_section = None
    footer_section = None
    if 'common_sections' in config:
        for section in config['common_sections']:
            if 'ヘッダー' in section['name']:
                header_section = section
            elif 'フッター' in section['name']:
                footer_section = section
    
    # 各ページの処理
    for page_idx, page in enumerate(config['pages']):
        page_name = page['name']
        
        # ページの切り替え時に空白行を追加（最初のページ以外）
        if page_idx > 0:
            current_row += 1
        
        # このページの開始行を記録
        page_row_start = current_row
        
        # 1. ヘッダー（共通部分）- トップページのみ
        if page_idx == 0 and header_section:
            content = header_section.get('content', '')
            if not content and 'elements' in header_section:
                content = '\n'.join([e.get('content', '') for e in header_section['elements']])
            add_section_row(page_name, header_section['name'], content)
        
        # 2. ページ固有のセクション
        for section in page.get('sections', []):
            section_name = section['name']
            content = section.get('content', '')
            if not content and 'elements' in section:
                content = '\n'.join([e.get('content', '') for e in section['elements']])
            add_section_row(page_name, section_name, content)
        
        # 3. フッター（共通部分）- トップページのみ
        if page_idx == 0 and footer_section:
            content = footer_section.get('content', '')
            if not content and 'elements' in footer_section:
                content = '\n'.join([e.get('content', '') for e in footer_section['elements']])
            add_section_row(page_name, footer_section['name'], content)
        
        # ページのセル統合（開始行から現在行の1つ前まで）
        page_row_end = current_row - 1
        if page_row_end > page_row_start:
            ws.merge_cells(start_row=page_row_start, start_column=1, 
                          end_row=page_row_end, end_column=1)
            first_cell = ws.cell(row=page_row_start, column=1)
            first_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
    
    # 空白行
    current_row += 2
    
    # 全体的な修正指示セクション
    ws.cell(row=current_row, column=1).value = "全体的な修正指示・その他のご要望"
    ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color="000000")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    overall_instruction = """以下の内容をこちらにご記入ください：
・サイト全体に関する修正指示
・新しいページやセクションの追加
・削除したいページやセクション
・その他、上記の表に当てはまらないご要望"""
    ws.cell(row=current_row, column=1).value = overall_instruction
    ws.cell(row=current_row, column=1).font = Font(size=11, color="000000")
    ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=4).border = thin_border
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.row_dimensions[current_row].height = 100
    
    # 列幅設定
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 50
    
    # 注意: 先頭行固定（freeze_panes）は適用しない
    
    # 保存
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    # サイト名を英数字に変換（日本語を除去）
    safe_site_name = config['site_name'].replace(' ', '_')
    filename = f"{safe_site_name}_site_guide_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    print(f"\n✓ 完成！")
    print(f"✓ ファイルを保存しました: {filepath}")
    
    return filepath


def main():
    """メイン処理"""
    import sys
    
    print("=" * 60)
    print("Website Revision Guide Generator v3.0")
    print("=" * 60)
    print()
    
    config_path = sys.argv[1] if len(sys.argv) > 1 else 'config.yaml'
    output_dir = sys.argv[2] if len(sys.argv) > 2 else 'output'
    
    filepath = generate_guide(config_path, output_dir)
    
    print()
    print("=" * 60)
    print("生成完了！")
    print("=" * 60)


if __name__ == '__main__':
    main()
